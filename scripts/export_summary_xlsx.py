"""
Export a two-tab Excel workbook summarising the dataset:
  Tab 1 (JA)   日本語版
  Tab 2 (EN)   English version

Output: pandora_database_summary.xlsx  (project root)
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PROJECT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT = os.path.join(PROJECT_DIR, "pandora_database_summary.xlsx")

# ---- Styling helpers ----
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FILL_SECTION = PatternFill("solid", fgColor="1F4E78")
FONT_SECTION = Font(name="Calibri", size=13, bold=True, color="FFFFFF")

FILL_SUB = PatternFill("solid", fgColor="D9E2F3")
FONT_SUB = Font(name="Calibri", size=11, bold=True, color="1F4E78")

FILL_HEAD = PatternFill("solid", fgColor="E7E6E6")
FONT_HEAD = Font(name="Calibri", size=10, bold=True)

FONT_BODY = Font(name="Calibri", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_section(ws, row, title, span):
    ws.cell(row=row, column=1, value=title).font = FONT_SECTION
    ws.cell(row=row, column=1).fill = FILL_SECTION
    ws.cell(row=row, column=1).alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 22
    if span > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    return row + 1


def write_sub(ws, row, title, span):
    ws.cell(row=row, column=1, value=title).font = FONT_SUB
    ws.cell(row=row, column=1).fill = FILL_SUB
    ws.cell(row=row, column=1).alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 18
    if span > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    return row + 1


def write_table(ws, row, headers, rows_data):
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = FONT_HEAD
        cell.fill = FILL_HEAD
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    row += 1
    for data in rows_data:
        for c, v in enumerate(data, start=1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.font = FONT_BODY
            cell.border = BORDER
            cell.alignment = WRAP
        row += 1
    return row + 1  # blank row after


# ============================================================
# JAPANESE SHEET
# ============================================================

def build_ja(ws):
    set_col_widths(ws, [28, 52, 20, 18])

    r = 1
    ws.cell(row=r, column=1, value="Pandora Paradox — データベース構成サマリー").font = Font(
        name="Calibri", size=16, bold=True, color="1F4E78"
    )
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws.row_dimensions[r].height = 26
    r += 1
    ws.cell(row=r, column=1, value="実データで構築したデータセット(197作品 × 87列)の概要").font = Font(
        name="Calibri", size=11, italic=True, color="595959"
    )
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    r += 2

    # ---- Section 1 ----
    r = write_section(ws, r, "1. 取得できた特徴量", 4)

    # 1a Metadata
    r = write_sub(ws, r, "メタデータ層(制作段階で既知)", 4)
    r = write_table(
        ws, r,
        ["カテゴリ", "特徴量", "取得元", "カバー率"],
        [
            ("基本情報", "rank, year, worldwide_gross, domestic_gross, budget",
             "top200_movies.csv", "100%"),
            ("分類情報", "genre, franchise, director, studio, is_sequel, is_animated",
             "top200_movies.csv", "100%"),
        ],
    )

    # 1b Y-side
    r = write_sub(ws, r, "文化的インパクト層(Y側 / CFI成分)", 4)
    r = write_table(
        ws, r,
        ["カテゴリ", "特徴量", "取得元", "カバー率"],
        [
            ("Wikipedia",
             "wiki_total_views, wiki_avg_monthly_views, wiki_max_monthly_views, wiki_languages",
             "Wikimedia API", "98%"),
            ("Reddit",
             "reddit_posts, reddit_comments, reddit_upvotes, reddit_top_post_upvotes",
             "Reddit public API", "99%"),
            ("Reddit派生",
             "meme_post_count, subreddit_diversity, subreddit_concentration",
             "Reddit posts派生", "100%"),
            ("TMDB",
             "tmdb_popularity, tmdb_vote_average, tmdb_vote_count, tmdb_runtime",
             "TMDB API", "98%"),
            ("OMDb / IMDb",
             "imdb_rating, imdb_votes, metascore, oscar_wins, total_wins, total_nominations",
             "OMDb API", "96%"),
            ("Wikiquote",
             "wikiquote_count (対話ブロック数)",
             "Wikiquoteスクレイピング", "93%"),
        ],
    )

    # 1c X-side
    r = write_sub(ws, r, "脚本・台詞特徴量層(X側 / 予測子)", 4)
    r = write_table(
        ws, r,
        ["カテゴリ", "特徴量", "取得元", "カバー率"],
        [
            ("基本統計",
             "dialogue_line_count, avg_dialogue_length, estimated_word_count, longest_monologue_words",
             "YIFY字幕", "95%"),
            ("感情",
             "sentiment_mean, sentiment_std, sentiment_peak, sentiment_spike",
             "字幕+辞書", "95%"),
            ("文体",
             "exclamation_ratio, question_ratio, profanity_score, vocabulary_richness, avg_sentence_complexity",
             "字幕", "95%"),
            ("主題",
             "humor_indicator, romance_indicator, violence_indicator",
             "字幕+辞書", "95%"),
            ("メーム予測用(新規設計)",
             "short_punchy_count, short_punchy_density, rare_proper_noun_count",
             "字幕", "95%"),
            ("脚本構造",
             "imsdb_dialogue_percentage, imsdb_total_lines",
             "IMSDb", "22%(補完用)"),
        ],
    )

    # 1d CFI
    r = write_sub(ws, r, "CFI合成指標", 4)
    r = write_table(
        ws, r,
        ["項目", "内容", "備考", ""],
        [
            ("CFI成分(8つ)",
             "cfi_comp_wiki_views, cfi_comp_wiki_langs, cfi_comp_reddit, cfi_comp_tmdb_votes, cfi_comp_gtrends, cfi_comp_quotes, cfi_comp_awards, cfi_comp_memes",
             "各成分を0-1に正規化", ""),
            ("重み(extended)",
             "20% Wiki views / 10% Wiki langs / 20% Reddit / 10% TMDB votes / 10% Trends / 10% Wikiquote / 10% Awards / 10% Memes",
             "合計100%", ""),
            ("合成スコア",
             "cultural_footprint_index → cfi_score (0-100にスケール)",
             "", ""),
            ("パラドックス指標",
             "cfi_residual (興収+予算の回帰からの残差), cultural_category",
             "正=オーバーパフォーマー、負=アンダーパフォーマー", ""),
        ],
    )

    # ---- Section 2 ----
    r = write_section(ws, r, "2. 取得できなかった / カバー率が低かったもの", 4)
    r = write_table(
        ws, r,
        ["指標", "カバー率", "原因", "影響度"],
        [
            ("Google Trends", "7% (13/197)",
             "Googleのレート制限(429エラー184件)。再試行には数時間〜数日の間隔が必要",
             "中 — CFIで10%を担う予定だったが実質機能せず"),
            ("IMSDb脚本", "22% (43/197)",
             "そもそも大半のブロックバスターがIMSDbに未登録",
             "低 — YIFY字幕95%でカバー済み"),
            ("IMDb直接アクセス(引用数)", "0%",
             "HTTP 202 anti-botブロック。Selenium等を使えば可能だがTOS違反リスク",
             "低 — Wikiquote 93%で代替実現"),
            ("Merchandise指数", "0%",
             "スケーラブルに取得できる公開データソースが存在しない",
             "中 — 元CFIに含まれていたが代替なく諦めた"),
            ("ファン活動規模(Fandom wiki等)", "0%",
             "未実装(次の候補)",
             "低"),
        ],
    )

    # ---- Section 3 ----
    r = write_section(ws, r, "3. 今後追加すると精度が上がりそうなもの", 4)

    # 3A
    r = write_sub(ws, r, "A. データ拡張系", 4)
    r = write_table(
        ws, r,
        ["候補", "理由", "難易度", ""],
        [
            ("サンプル拡張(カルト作品を含める)",
             "現在Top200興収限定 → 低興収・高文化作品(The Room等)が除外されており選択バイアスが大きい",
             "★☆☆", ""),
            ("years_since_release を明示化",
             "古い作品ほどWikiビューが蓄積 → 時間交絡の制御",
             "★☆☆", ""),
            ("release_year コホート化",
             "世代効果(8-15歳で観た作品が後年に過剰評価される)の制御",
             "★☆☆", ""),
        ],
    )

    # 3B
    r = write_sub(ws, r, "B. 予測力を上げそうな新特徴量", 4)
    r = write_table(
        ws, r,
        ["候補", "情報の種類", "期待される効果", ""],
        [
            ("YouTube予告編指標(再生数/コメント/いいね)",
             "公開前バズ",
             "事前予測(Model A)の主要因子候補", ""),
            ("キャスト/監督のスターパワー(過去作CFI, フォロワー数)",
             "人的資本",
             "監督/主演バイアスの定量化", ""),
            ("原作IPの強度(原作書籍部数, コミックシリーズ期間)",
             "既存ファンベース",
             "フランチャイズ効果の源泉を直接測定", ""),
            ("マーケティング予算(製作予算とは別)",
             "宣伝投資",
             "現状 budget と切り分けできていない", ""),
            ("配信プラットフォーム履歴(Netflix/Disney+での視聴期間)",
             "長期接触機会",
             "時間経過で記憶が維持される仕組みを直接測定", ""),
            ("サウンドトラック識別性(Shazam/Spotify再生数)",
             "音による記憶",
             "台詞では捉えられない音響的メーム", ""),
            ("予告編/ポスターの視覚独自性",
             "視覚的記憶",
             "「ジョーカー階段」的な1フレーム記憶", ""),
        ],
    )

    # 3C
    r = write_sub(ws, r, "C. モデリング改善系", 4)
    r = write_table(
        ws, r,
        ["候補", "内容", "", ""],
        [
            ("時間次元の特徴量化",
             "Wiki/Redditの月別データを時系列特徴量に展開", "", ""),
            ("LLMによる引用抽出",
             "字幕テキストから「メーム化しそうな台詞」をLLMで同定 → カウント", "", ""),
            ("CFI重み付けの学習",
             "現状ヒューリスティック10〜20% → PCA/因子分析で自動決定", "", ""),
            ("交互作用項",
             "franchise × short_punchy_count 等の掛け合わせ効果", "", ""),
        ],
    )

    # 3D
    r = write_sub(ws, r, "D. 最優先で入れたい3つ", 4)
    r = write_table(
        ws, r,
        ["優先順位", "項目", "理由", ""],
        [
            ("1", "years_since_release の追加",
             "10分で追加、時間交絡を制御して全モデル即時改善の可能性", ""),
            ("2", "YouTube予告編再生数",
             "YouTube Data APIで無料、ほぼ全作品で取得可能、事前予測の主力候補", ""),
            ("3", "サンプル拡張(カルト作品30〜50本追加)",
             "選択バイアスを減らし、パラドックスの核心ケースを分析可能に", ""),
        ],
    )


# ============================================================
# ENGLISH SHEET
# ============================================================

def build_en(ws):
    set_col_widths(ws, [32, 56, 22, 22])

    r = 1
    ws.cell(row=r, column=1, value="Pandora Paradox — Database Summary").font = Font(
        name="Calibri", size=16, bold=True, color="1F4E78"
    )
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws.row_dimensions[r].height = 26
    r += 1
    ws.cell(row=r, column=1, value="Overview of the real-data dataset built from 197 films × 87 columns").font = Font(
        name="Calibri", size=11, italic=True, color="595959"
    )
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    r += 2

    # Section 1
    r = write_section(ws, r, "1. Collected features", 4)

    r = write_sub(ws, r, "Metadata layer (known at production stage)", 4)
    r = write_table(
        ws, r,
        ["Category", "Features", "Source", "Coverage"],
        [
            ("Basic", "rank, year, worldwide_gross, domestic_gross, budget",
             "top200_movies.csv", "100%"),
            ("Classification", "genre, franchise, director, studio, is_sequel, is_animated",
             "top200_movies.csv", "100%"),
        ],
    )

    r = write_sub(ws, r, "Cultural impact layer (Y-side / CFI components)", 4)
    r = write_table(
        ws, r,
        ["Category", "Features", "Source", "Coverage"],
        [
            ("Wikipedia",
             "wiki_total_views, wiki_avg_monthly_views, wiki_max_monthly_views, wiki_languages",
             "Wikimedia API", "98%"),
            ("Reddit",
             "reddit_posts, reddit_comments, reddit_upvotes, reddit_top_post_upvotes",
             "Reddit public API", "99%"),
            ("Reddit-derived",
             "meme_post_count, subreddit_diversity, subreddit_concentration",
             "Derived from Reddit posts", "100%"),
            ("TMDB",
             "tmdb_popularity, tmdb_vote_average, tmdb_vote_count, tmdb_runtime",
             "TMDB API", "98%"),
            ("OMDb / IMDb",
             "imdb_rating, imdb_votes, metascore, oscar_wins, total_wins, total_nominations",
             "OMDb API", "96%"),
            ("Wikiquote",
             "wikiquote_count (dialogue block count)",
             "Wikiquote scraping", "93%"),
        ],
    )

    r = write_sub(ws, r, "Script / dialogue feature layer (X-side / predictors)", 4)
    r = write_table(
        ws, r,
        ["Category", "Features", "Source", "Coverage"],
        [
            ("Basic stats",
             "dialogue_line_count, avg_dialogue_length, estimated_word_count, longest_monologue_words",
             "YIFY subtitles", "95%"),
            ("Sentiment",
             "sentiment_mean, sentiment_std, sentiment_peak, sentiment_spike",
             "Subtitles + lexicon", "95%"),
            ("Style",
             "exclamation_ratio, question_ratio, profanity_score, vocabulary_richness, avg_sentence_complexity",
             "Subtitles", "95%"),
            ("Theme",
             "humor_indicator, romance_indicator, violence_indicator",
             "Subtitles + lexicon", "95%"),
            ("Meme-prediction (newly designed)",
             "short_punchy_count, short_punchy_density, rare_proper_noun_count",
             "Subtitles", "95%"),
            ("Script structure",
             "imsdb_dialogue_percentage, imsdb_total_lines",
             "IMSDb", "22% (fallback)"),
        ],
    )

    r = write_sub(ws, r, "CFI composite index", 4)
    r = write_table(
        ws, r,
        ["Item", "Details", "Note", ""],
        [
            ("CFI components (8)",
             "cfi_comp_wiki_views, cfi_comp_wiki_langs, cfi_comp_reddit, cfi_comp_tmdb_votes, cfi_comp_gtrends, cfi_comp_quotes, cfi_comp_awards, cfi_comp_memes",
             "Each normalised to 0-1", ""),
            ("Weights (extended)",
             "20% Wiki views / 10% Wiki langs / 20% Reddit / 10% TMDB votes / 10% Trends / 10% Wikiquote / 10% Awards / 10% Memes",
             "Sums to 100%", ""),
            ("Composite score",
             "cultural_footprint_index -> cfi_score (scaled 0-100)",
             "", ""),
            ("Paradox metric",
             "cfi_residual (residual from box-office + budget regression), cultural_category",
             "Positive = overperformer, negative = underperformer", ""),
        ],
    )

    # Section 2
    r = write_section(ws, r, "2. Data not collected / weakly covered", 4)
    r = write_table(
        ws, r,
        ["Metric", "Coverage", "Cause", "Impact"],
        [
            ("Google Trends", "7% (13/197)",
             "Google rate-limited us hard (184 HTTP 429 errors). Retrying needs hours-to-days gaps.",
             "Medium — intended to carry 10% of CFI but effectively non-functional"),
            ("IMSDb scripts", "22% (43/197)",
             "Most blockbusters aren't listed on IMSDb in the first place",
             "Low — YIFY subtitles cover 95%"),
            ("Direct IMDb access (quotes)", "0%",
             "HTTP 202 anti-bot block. Possible via Selenium but carries TOS risk.",
             "Low — substituted with Wikiquote (93%)"),
            ("Merchandise index", "0%",
             "No scalable public data source exists",
             "Medium — was in original CFI but dropped due to no substitute"),
            ("Fandom activity (Fandom wiki etc.)", "0%",
             "Not yet implemented (next candidate)",
             "Low"),
        ],
    )

    # Section 3
    r = write_section(ws, r, "3. Additions likely to improve the model", 4)

    r = write_sub(ws, r, "A. Data expansion", 4)
    r = write_table(
        ws, r,
        ["Candidate", "Rationale", "Difficulty", ""],
        [
            ("Sample expansion (include cult films)",
             "Currently limited to Top 200 box office. Low-gross / high-culture films (e.g. The Room) are excluded, introducing a large selection bias.",
             "★☆☆", ""),
            ("Add years_since_release",
             "Older films accumulate more Wiki views — control the time confound.",
             "★☆☆", ""),
            ("release_year cohort",
             "Control generational effects (films seen at age 8-15 tend to be over-valued later).",
             "★☆☆", ""),
        ],
    )

    r = write_sub(ws, r, "B. New features likely to improve prediction", 4)
    r = write_table(
        ws, r,
        ["Candidate", "Information type", "Expected effect", ""],
        [
            ("YouTube trailer metrics (views/comments/likes)",
             "Pre-release buzz",
             "Strong candidate for Model A main driver", ""),
            ("Cast/director star power (prior CFI, follower counts)",
             "Human capital",
             "Quantify director/lead bias", ""),
            ("Source IP strength (book sales, comic run length)",
             "Existing fan base",
             "Directly measures franchise-effect source", ""),
            ("Marketing budget (separate from production budget)",
             "Promotional spend",
             "Currently confounded with budget", ""),
            ("Streaming availability history (Netflix/Disney+ duration)",
             "Long-term exposure",
             "Directly measures how memory is maintained over time", ""),
            ("Soundtrack distinctiveness (Shazam/Spotify play counts)",
             "Auditory memory",
             "Captures acoustic memes that dialogue doesn't", ""),
            ("Visual distinctiveness of trailer/poster",
             "Visual memory",
             "Captures single-frame memories like the Joker stairs scene", ""),
        ],
    )

    r = write_sub(ws, r, "C. Modeling improvements", 4)
    r = write_table(
        ws, r,
        ["Candidate", "Details", "", ""],
        [
            ("Time-dimension featurisation",
             "Expand monthly Wiki/Reddit data into time-series features", "", ""),
            ("LLM-based quote extraction",
             "Use an LLM to identify likely-memeable lines in subtitle text, then count them", "", ""),
            ("Learned CFI weights",
             "Replace current 10-20% heuristic weights with PCA / factor-analysis derived weights", "", ""),
            ("Interaction terms",
             "e.g. franchise × short_punchy_count", "", ""),
        ],
    )

    r = write_sub(ws, r, "D. Top 3 priorities to add", 4)
    r = write_table(
        ws, r,
        ["Priority", "Item", "Rationale", ""],
        [
            ("1", "Add years_since_release",
             "10-minute change; controls time confound and likely lifts all models immediately", ""),
            ("2", "YouTube trailer view counts",
             "Free via YouTube Data API; available for almost all films; strongest Model A candidate", ""),
            ("3", "Sample expansion (+30-50 cult films)",
             "Reduces selection bias and allows direct study of paradox-core cases", ""),
        ],
    )


# ============================================================
# Main
# ============================================================

def main():
    wb = Workbook()
    ws_ja = wb.active
    ws_ja.title = "日本語"
    build_ja(ws_ja)

    ws_en = wb.create_sheet("English")
    build_en(ws_en)

    wb.save(OUTPUT)
    print(f"Wrote: {OUTPUT}")


if __name__ == "__main__":
    main()
